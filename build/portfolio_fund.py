# -*- coding: utf-8 -*-
r"""build/portfolio_fund.py — 운용 포트폴리오 페이지(portfolio.html)의 본문 조각을 만든다.

무엇을 만드나. 실펀드 2종(2Z30=나스닥100 · 2A81=S&P500)의
  ① 펀드 개요 — NAV·기준가, 연초 후 기준가 vs 지수(원화환산), 자산 구성
  ② 보유 vs 지수 — 종목별 지수비중/펀드비중/액티브 틸트, 패시브 대비 수량 괴리
  ③ 전략 매매 내역 — 웹 원장(+이전 전이면 DB 씨앗) + 현재가 평가
  ④ 전략 성과·기여 — 매매 시점 일치 BM 대비 초과손익과 NAV 기여(bp)
를 렌더 완료된 HTML 조각으로 _build/pages/portfolio_content.html 에 쓴다.

🚨 이 조각은 **평문 그대로 저장소에 들어가지 않는다.** _build/ 는 gitignore 이고,
   배포는 build/kb_lock.py --page portfolio 가 AES-256-GCM 으로 잠가 portfolio.html 의
   PAYLOAD 에 넣는다(kb.html 과 같은 규약 — 열람 암호 없이는 소스를 봐도 못 푼다).
   실펀드 NAV·보유 수량·매매 내역이라 공개 사이트에 평문으로 나가면 안 된다.

입력 셋 — 전부 이 사내 PC 에서만 접근 가능하다(러너 이식 불가, 로컬 수동 잡).
  · 사내 export(엑셀): 경로는 _build/portfolio_local.json 의 xlsm_globs (gitignore — 내부
      파일서버 경로라 공개 저장소에 안 적는다). NAV·환율·해외(보유 원장) 시트를 읽는다.
      ⚠ 시트 이름·컬럼이 사내 시스템 export 규격이다. 바뀌면 여기가 아니라 규격이 바뀐 것.
  · 사내 DB — 자격증명·호스트는 연구 repo util/variables.py 가 단일 출처다(아래 _db_params.
      🚨 이 저장소는 공개라 여기에 절대 적지 않는다 — 2026-08-20 적대감사가 평문 노출을 잡았다).
      **public.index_constituents(지수 비중·GICS) 하나만 쓴다**(2026-08-21 사용자 결정).
      나머지 셋은 이 랩의 웹 자료로 옮겼다 — 아래 «자료 원천» 참조.
  · 이 랩의 웹 자료(yfinance · CI 가 매일 굽는다 · 사내망 불필요):
      data/sd/<티커>.json  종목 종가(분할 소급 조정) · data/assets.json  ^GSPC·^NDX 지수 레벨
      data/splits.json     분할 이력 — 원장 체결가를 오늘 기준으로 되맞추는 데 쓴다
  · 웹 원장 data/portfolio_user.json — 전략 매매(AES-256-GCM). mp.strategy_trade 는
      이전용 씨앗으로만 한 번 더 읽고, 화면에서 옮기면 그 뒤로는 안 본다.

자산구분 코드(해외 시트, 실측 2026-08-18): 1=개별주식 · 3=지수 ETF · 4=지수선물(평가액=노셔널)
  · 5=예금 · B=증거금. 선물 노셔널은 NAV 에 없고 노출에만 더한다 — 섞으면 합이 100%를 넘는다.

정의(화면 각주와 같아야 한다 — 렌더가 이 문서를 그대로 옮긴다):
  · 시점 규약(T-1, 실측 검증 2026-08-20): 기준가(D) = 미국 D−1(직전 거래일) 종가 × D일
    한국마감 환율. NAV·환율은 보유일로 자르지 않고 시트 최신 행을 쓰고, 연초 후 차트는
    지수를 last_lt(하루 밀기)로 짝 맞춘다. 같은날 짝은 벤치가 하루 앞서 달리는 오류다.
  · 펀드비중 = 종목 평가액(원화) ÷ 개별주식 슬리브 합. 지수비중과 같은 눈금이 되도록
    주식 슬리브 안에서 정규화한다(펀드는 주식+ETF+선물로 지수를 복제하므로 NAV 분모로는
    전 종목이 일괄 언더웨이트로 보인다 — 그건 틸트가 아니라 구조다).
  · 패시브 수량 = 지수비중 × 주식슬리브(원) ÷ (종가 × USD환율). 괴리 = 실제 − 패시브.
  · 전략 수익률 = Σ수량×(현재가−체결가) ÷ 매수원금. 체결가는 원장의 trade_price 로,
    **당시 종가이지 실제 체결가가 아니다**(MP 엑셀 VBA 가 종가를 박는다).
  · BM 대조 = 같은 날 같은 금액을 지수에 넣었을 때의 손익(매매 시점 일치). 초과 = 차이.
  · NAV 기여(bp) = 초과손익(USD) × 기준일 환율 ÷ 기준일 NAV × 10,000.

한계(화면에 싣는다):
  · 배당 미반영 — 종가는 배당을 안 담는다. 보유 2개월 내외라 왜곡은 작지만 0이 아니다.
  · 분할은 **소급 조정된 랩 종가**를 쓰고, 원장 체결가는 split_factor() 로 같은 눈금에
    맞춘다(가격÷f · 수량×f — 금액 불변). 종전 사내 종가는 분할을 소급 안 해서 그 자체가
    틀린 값이었다(사용자 지적 2026-08-21 · 랩 실측 사고 MNST 2026-08-11).
  · 분할 가드(일수익 |40%| 초과 ⚠)는 남긴다 — 되맞춤이 빗나가는 경우를 잡는 이중 방어다.
  · 지수 밖 보유는 랩 유니버스에 없어 가격이 빠진다. 사용자 판단: «편출 예정이거나 합병
    등으로 일시적으로 생긴 것이라 가격 없어도 상관없음».

사용:  python build/portfolio_fund.py            # 조각 생성
       python build/kb_lock.py --page portfolio  # 잠가서 portfolio.html 에 기록(암호 입력)
"""
from __future__ import annotations

import datetime as dt
import glob
import html
import io
import json
import os
import sys

try: sys.stdout.reconfigure(encoding="utf-8")   # Windows 콘솔(cp949)에서 ⚠·— 출력 시 UnicodeEncodeError 방지
except Exception: pass

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")          # 랩 웹 자료(종가·지수·분할) — 사내 DB 대체분
OUT = os.path.join(ROOT, "_build", "pages", "portfolio_content.html")
LOCAL_CFG = os.path.join(ROOT, "_build", "portfolio_local.json")


def _local_cfg():
    """gitignore 된 로컬 설정(_build/portfolio_local.json) — 내부 파일서버 경로 등
    공개 저장소에 적으면 안 되는 값을 담는다. xlsm_globs 는 슬래시형 UNC 로 적을 것
    (역슬래시형은 편집 도구를 거치며 이스케이프가 깨진 전력이 있다 — 2026-08-20 실측)."""
    if not os.path.exists(LOCAL_CFG):
        raise SystemExit("로컬 설정 없음: %s — {\"xlsm_globs\": [...]} 형태로 만들 것"
                         "(내부 경로라 저장소에 안 들어간다)" % LOCAL_CFG)
    return json.load(io.open(LOCAL_CFG, encoding="utf-8"))


def _db_params():
    """DB 자격증명 — 연구 repo util/variables.py 가 단일 출처(호스트 포함). 이 파일에 상수로
    두지 않는다: 이 저장소는 공개이고, 실제로 평문 노출 사고가 있었다(2026-08-20 적대감사,
    그 전 이력에는 남아 있어 별도 조치 대상). build/db_load.py 와 같은 사유·비슷한 순서다."""
    repo = os.path.expanduser(os.getenv("YEOUIDO_REPO") or "C:/Projects/Yeouido")
    if os.path.exists(os.path.join(repo, "util", "variables.py")):
        sys.path.insert(0, repo)
        try:
            import util.variables as V
            return dict(host=V.host, port=5432, dbname=V.database,
                        user=V.user, password=V.password, connect_timeout=12)
        finally:
            sys.path.remove(repo)
    env = {k: os.environ.get("YEOUIDO_DB_" + k) for k in ("HOST", "PORT", "NAME", "USER", "PASS")}
    if env["HOST"] and env["USER"] and env["PASS"]:
        return dict(host=env["HOST"], port=int(env["PORT"] or 5432),
                    dbname=env["NAME"] or "postgres", user=env["USER"],
                    password=env["PASS"], connect_timeout=12)
    raise SystemExit("DB 접속 정보 없음 — 연구 repo util/variables.py(YEOUIDO_REPO) 또는 "
                     "환경변수 YEOUIDO_DB_HOST/USER/PASS 를 준비할 것")

# 탭 순서 = 이 목록 순서. 2026-08-20 사용자 지시 — «S&P500을 왼쪽에, 나스닥100을 오른쪽에».
FUNDS = [
    ("2A81", "SPX Index", "spx", "S&P500"),
    ("2Z30", "NDX Index", "ndx", "나스닥100"),
]
SPLIT_GUARD = 0.40      # 매매 구간 일수익 절대값이 이걸 넘으면 분할 의심 ⚠


def esc(s):
    return html.escape(str(s if s is not None else ""), quote=True)


def num(v, nd=2):
    """천 단위 구분 표기. None 은 —."""
    if v is None:
        return "—"
    if nd == 0:
        return format(int(round(float(v))), ",")
    return format(round(float(v), nd), ",.%df" % nd)


def pct(v, nd=2, signed=False):
    if v is None:
        return "—"
    s = ("%+." if signed else "%.") + str(nd) + "f"
    return (s % (v * 100)) + "%"


def cls_sign(v):
    return "pos" if (v or 0) > 0 else ("neg" if (v or 0) < 0 else "")


def norm_tk(t):
    """티커 정규화 — 클래스주 구분자를 점으로 통일한다(BRK/B → BRK.B · BF/B → BF.B).

    🚨 2026-08-20 사용자 발견: 펀드가 들고 있는 BRK/B(사내 시트 · 블룸버그식 슬래시)가
      지수 구성종목의 BRK.B(팩트셋식 점)와 안 맞아, 보유 중인데 «지수에 있는데 미보유»
      목록에 나오고 표의 지수비중은 0 이었다. BF/B 도 같다.
    모든 입구(시트·지수구성·매매원장·종가)가 이 함수를 지나므로 어느 원천이 어느 표기를
    쓰든 안에서는 한 이름이다. 종가 조회만은 DB 표기를 모르니 **두 표기를 다 물어보고**
    돌아온 것을 정규화해 담는다(load_db 참조)."""
    return (t or "").strip().replace("/", ".")


def lab_tk(t):
    """사내 표기 → 이 랩의 종목 파일 이름. «NVDA US» → «NVDA» · «BRK/B US» → «BRK.B».

    랩은 data/sd/<티커>.json 이고 클래스주를 점으로 쓴다(실측: BRK.B.json · BF.B.json) —
    norm_tk 가 이미 슬래시를 점으로 바꿔 두므로 여기서는 거래소 접미사만 뗀다.
    ⚠ 미국 상장분만 뗀다. 다른 접미사(« LN» 등)는 랩 유니버스에 없어 어차피 못 찾는다."""
    t = norm_tk(t)
    return t[:-3] if t.endswith(" US") else t


# ── 1) 사내 export 읽기 ──────────────────────────────────────────────────────
def load_xlsm():
    # mtime 동률(같은 저장을 양쪽에 복사한 경우)이면 네트워크 정본을 이긴다
    files = sorted((f for g in _local_cfg()["xlsm_globs"] for f in glob.glob(g)),
                   key=lambda f: (os.path.getmtime(f), f.replace(chr(92), '/').startswith('//')))
    if not files:
        raise SystemExit("사내 export 없음 — 네트워크 공유((주간) 자동화 2Z30*.xlsm) 또는 " "SecureGate 다운로드에 파일을 둘 것")
    path = files[-1]
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    def d10(v):
        if isinstance(v, dt.datetime):
            return v.strftime("%Y-%m-%d")
        return str(v)[:10]

    nav = {}          # fund → {date: (nav, base_price)}
    for r in wb["NAV"].iter_rows(min_row=11, values_only=True):
        if not r[0]:
            break
        nav.setdefault(str(r[1]).strip(), {})[d10(r[0])] = (float(r[2]), float(r[4]))

    fx = {}           # date → USD 환율(종가 기준)
    for r in wb["환율"].iter_rows(min_row=11, values_only=True):
        if not r[0]:
            break
        if str(r[1]).strip() == "USD":
            fx[d10(r[0])] = float(r[2])

    hold = {}         # fund → {date: [보유행…]}
    for r in wb["해외"].iter_rows(min_row=11, values_only=True):
        if not r[0]:
            break
        hold.setdefault(str(r[0]).strip(), {}).setdefault(d10(r[4]), []).append({
            "ticker": (norm_tk(str(r[2])) if r[2] else ""),
            "name": (str(r[3]).strip() if r[3] else ""),
            "qty": float(r[7] or 0), "val_usd": float(r[10] or 0),
            "val_krw": float(r[11] or 0), "asset": str(r[13]).strip(),
            "px": (float(r[18]) if r[18] not in (None, "") else None),
        })
    return path, nav, fx, hold


# ── 2) 사내 DB 읽기 ──────────────────────────────────────────────────────────
PANEL_DAYS = 504    # 웹 앱에 동봉하는 종가 패널 길이(거래일 ≈ 2년). 웹 백테스트·성과 재계산의 창.


def load_db(asof_by_fund, held_tickers):
    """지수 구성종목(비중·이름·GICS) — **이것만 사내 DB에서 온다.**

    🚨 2026-08-21 사용자 결정으로 나머지 셋은 웹(yfinance) 자료로 옮겼다:
        market.ohlcv_factset      → data/sd/<티커>.json      (랩이 매일 굽는 종가)
        public.price_major_index  → data/assets.json         (^GSPC · ^NDX)
        mp.strategy_trade         → 웹 원장(data/portfolio_user.json)
      사용자 판단 근거를 그대로 적어 둔다 —
        · 지수 밖 보유는 «편출 예정이거나 합병 등으로 일시적으로 생긴 것»이라 가격이 없어도 된다.
        · **사내 원장의 가격이 오히려 틀리다** — 분할을 소급 반영하지 않는다. 랩 자료는
          분할 소급이 검사로 보장된다(validate_site 의 «분할 소급 검사»).
      지수 비중만은 웹에 대체 자료가 없다(랩은 시총·유동주식수를 안 들고 있다). 그래서 남긴다.
    ⚠ mp.strategy_trade 는 **이전용으로만** 한 번 더 읽는다(아래 seed) — 웹 원장으로 옮기고
      나면 이 읽기도 지운다. 옮기기 전까지 화면이 비면 안 되니까 남겨 둔 다리다.
    """
    import psycopg2
    cn = psycopg2.connect(**_db_params())
    cur = cn.cursor()

    cons = {}         # index → (dt, {ticker: (정규화 비중, 이름, GICS)})
    for _f, idx, _s, _l in FUNDS:
        cur.execute('SELECT max(dt) FROM public.index_constituents WHERE "index"=%s AND dt<=%s',
                    (idx, asof_by_fund[_f]))
        d = cur.fetchone()[0]
        cur.execute('SELECT ticker, index_weight, name, gics_name FROM public.index_constituents '
                    'WHERE "index"=%s AND dt=%s', (idx, d))
        rows = cur.fetchall()
        tot = sum(float(w or 0) for _t, w, _n, _g in rows) or 1.0
        cons[idx] = (str(d), {norm_tk(t): (float(w or 0) / tot, n, g) for t, w, n, g in rows})

    # 이전용 씨앗 — 웹 원장으로 한 번 옮기고 나면 이 블록을 지운다.
    seed = []
    try:
        cur.execute('SELECT "index", dt, strategy, ticker, trade_qty, trade_price '
                    'FROM mp.strategy_trade ORDER BY dt, strategy, ticker')
        seed = [dict(index=i, dt=str(d), strategy=s2, ticker=norm_tk(t),
                     qty=float(q), px=float(p or 0)) for i, d, s2, t, q, p in cur.fetchall()]
    except Exception as _e:
        print("  ⚠ mp.strategy_trade 를 못 읽었다(%s) — 웹 원장만 쓴다." % str(_e)[:60])
    cn.close()
    return cons, seed


def load_web(uni_tickers):
    """종가·지수 레벨을 **랩의 웹 자료**에서 읽는다(사내 DB 대체).

    · 거래일 축 = data/stocks.json 의 pxd_dates (랩 정본 격자 · 패딩 없음)
    · 종목 종가 = data/sd/<티커>.json 의 pxd (그 격자에 정렬돼 있다)
    · 지수 레벨 = data/assets.json 의 ^GSPC · ^NDX
    ⚠ 랩 종가는 **분할 소급 조정**됐다. 사내 원장의 체결가는 무조정이라, 매매 이후 분할한
      종목은 둘을 그대로 빼면 안 된다 — split_factor() 가 원장 체결가를 오늘 기준으로
      되맞춘다(그쪽 주석 참조). 이 차이가 이 교체의 유일한 위험 지점이다.
    """
    st = json.load(io.open(os.path.join(DATA, "stocks.json"), encoding="utf-8"))
    grid = st["pxd_dates"]
    axis = grid[-PANEL_DAYS:]
    gi = {d: i for i, d in enumerate(grid)}

    px, miss = {}, []
    for t in sorted(uni_tickers):
        f = os.path.join(DATA, "sd", lab_tk(t) + ".json")
        if not os.path.exists(f):
            miss.append(t)
            continue
        arr = (json.load(io.open(f, encoding="utf-8")) or {}).get("pxd") or []
        ser = {}
        for d in axis:
            k = gi.get(d)
            if k is not None and k < len(arr) and arr[k]:
                ser[d] = float(arr[k])
        if ser:
            px[t] = ser

    A = json.load(io.open(os.path.join(DATA, "assets.json"), encoding="utf-8"))
    ad, ap = A["dates"], A["px"]
    lvl = {}
    for idx_name, sym in (("SPX Index", "^GSPC"), ("NDX Index", "^NDX")):
        arr = ap.get(sym) or []
        lvl[idx_name] = {d: float(v) for d, v in zip(ad, arr) if v}
    for idx_name in ("SPX Index", "NDX Index"):
        if not lvl.get(idx_name):
            raise SystemExit("지수 레벨을 못 읽었다(%s) — data/assets.json 을 먼저 갱신할 것" % idx_name)
    print("  웹 자료: 종가 %d종(축 %d일 %s~%s) · 지수 2종 · 못 찾은 티커 %d종%s"
          % (len(px), len(axis), axis[0], axis[-1], len(miss),
             (" (" + ", ".join(miss[:8]) + (" …" if len(miss) > 8 else "") + ")") if miss else ""))
    return px, lvl, axis


def split_factor(t, trade_dt):
    """체결일 이후의 분할비 곱 — 원장 체결가·수량을 **오늘 기준**으로 되맞추는 계수.

    🚨 왜 필요한가. 랩 종가는 분할 소급 조정본이고 사내 원장의 체결가는 그날의 무조정
      가격이다. 그대로 빼면 2:1 분할 종목의 수익률이 −50%로 찍힌다 — 이 랩이 실제로
      겪은 사고다(MNST 2026-08-11). 체결가 ÷ f · 수량 × f 로 되맞춘다(금액은 불변).
    반환 f: 체결일 «이후»에 일어난 분할들의 곱. 없으면 1.0.
    """
    ev = (_SPLITS or {}).get(lab_tk(t)) or []
    f = 1.0
    for d, r in ev:
        if d > trade_dt and r:
            f *= float(r)
    return f


_SPLITS = None


def load_splits():
    global _SPLITS
    try:
        _SPLITS = (json.load(io.open(os.path.join(DATA, "splits.json"), encoding="utf-8")) or {}).get("co") or {}
    except Exception:
        _SPLITS = {}
    return _SPLITS


def encode_panel(px, axis):
    """종가 패널 → 웹 앱용 압축 인코딩.

    한 종목의 종가를 «첫 유효 종가 p0 대비 비율 × 종목별 scale» 의 uint16 으로 담는다.
    0 은 결측 예약. scale 은 65000/최대비율로 종목마다 정하므로 2년에 10배 오른 종목도
    포화 없이 담기고, 해상도는 최악 0.015% — 성과·백테스트 눈금(1bp 단위)에 충분하다.
    float32 대비 절반 크기다 — 이 패널은 AES 암호문에 통째로 들어가 페이지 무게가 된다.
    바이트 순서는 리틀엔디언(x86 tobytes) — JS 의 Uint16Array 디코드와 같다.
    """
    import base64
    from array import array
    idx_of = {d: i for i, d in enumerate(axis)}
    # v>0 필터 — 벤더 0.0 종가가 섞이면 p0=0 으로 ZeroDivisionError(빌드 사망) 또는 결측
    # 예약값(0)과 충돌한다(적대감사 16). 0 종가는 가격이 아니라 결측이다.
    tickers = sorted(t for t, ser in px.items()
                     if any(d in idx_of and v > 0 for d, v in ser.items()))
    nd = len(axis)
    buf = array("H", bytes(2 * len(tickers) * nd))
    p0s, scales = [], []
    for k, t in enumerate(tickers):
        ser = [(idx_of[d], v) for d, v in px[t].items() if d in idx_of and v > 0]
        ser.sort()
        p0 = ser[0][1]
        mx = max(v for _i, v in ser)
        sc = min(8000.0, 65000.0 * p0 / mx)
        p0s.append(round(p0, 4))
        scales.append(round(sc, 4))
        base = k * nd
        for i, v in ser:
            buf[base + i] = max(1, min(65535, int(round(v / p0 * sc))))
    return {"tickers": tickers, "p0": p0s, "scale": scales, "nd": nd,
            "u16": base64.b64encode(buf.tobytes()).decode()}


def last_leq(series_dict, date):
    """{date: v} 에서 date 이하 마지막 (date, v). 없으면 (None, None)."""
    best = None
    for d in series_dict:
        if d <= date and (best is None or d > best):
            best = d
    return (best, series_dict[best]) if best else (None, None)


def last_lt(series_dict, date):
    """date «미만» 마지막 (date, v) — T-1 짝맞춤용. 기준가(D)는 직전 미국 거래일 종가를 담는다."""
    best = None
    for d in series_dict:
        if d < date and (best is None or d > best):
            best = d
    return (best, series_dict[best]) if best else (None, None)


# ── 3) 전략 성과 계산 ────────────────────────────────────────────────────────
def strat_perf(trades, px, lvl_idx, asof_us):
    """전략 → {curve, last, rows, warn_split}. BM 은 매매 시점 일치 지수 투자.

    손익 = Σ qty×(px_t − 체결가). 매도(qty<0)도 같은 식이 성립한다 — 매도 시점에
    잠근 손익이 이후 가격변동과 상쇄되어 실현분으로 남는다.
    """
    out = {}
    dates = sorted(d for d in lvl_idx if d <= asof_us)
    for tr in trades:
        s = out.setdefault(tr["strategy"], {"trades": [], "tickers": {}})
        s["trades"].append(tr)
        s["tickers"].setdefault(tr["ticker"], []).append(tr)

    for _sname, s in out.items():
        t0 = min(t["dt"] for t in s["trades"])
        curve = []
        for d in [x for x in dates if x >= t0]:
            pnl = bm = inv = 0.0
            for tr in s["trades"]:
                if tr["dt"] > d:
                    continue
                _pd, p = last_leq(px.get(tr["ticker"], {}), d)
                if p is None:
                    continue
                pnl += tr["qty"] * (p - tr["px"])
                i_t = lvl_idx.get(d) or last_leq(lvl_idx, d)[1]
                i_0 = lvl_idx.get(tr["dt"]) or last_leq(lvl_idx, tr["dt"])[1]
                if i_t and i_0:
                    bm += tr["qty"] * tr["px"] * (i_t / i_0 - 1)
                if tr["qty"] > 0:
                    inv += tr["qty"] * tr["px"]
            curve.append((d, pnl, bm, inv))
        s["curve"] = curve

        # 분할 가드 — 보유 구간에 하루 ±40% 를 넘는 변동이 있으면 의심 표기
        warn = set()
        for tk, trs in s["tickers"].items():
            t_first = min(t["dt"] for t in trs)
            rows = sorted((d, p) for d, p in px.get(tk, {}).items() if t_first <= d <= asof_us)
            for (d1, p1), (d2, p2) in zip(rows, rows[1:]):
                if p1 and abs(p2 / p1 - 1) > SPLIT_GUARD:
                    warn.add(tk)
        s["warn_split"] = sorted(warn)

        if curve:
            d, pnl, bm, inv = curve[-1]
            s["last"] = dict(dt=d, pnl=pnl, bm=bm, inv=inv,
                             ret=(pnl / inv if inv else None), bm_ret=(bm / inv if inv else None))

        rows = []
        for tk, trs in sorted(s["tickers"].items()):
            _pd, p = last_leq(px.get(tk, {}), asof_us)
            if p is None:
                continue
            inv_t = sum(t["qty"] * t["px"] for t in trs if t["qty"] > 0)
            pnl_t = sum(t["qty"] * (p - t["px"]) for t in trs)
            bm_t = 0.0
            i_t = last_leq(lvl_idx, asof_us)[1]
            for t in trs:
                i_0 = lvl_idx.get(t["dt"]) or last_leq(lvl_idx, t["dt"])[1]
                if i_t and i_0:
                    bm_t += t["qty"] * t["px"] * (i_t / i_0 - 1)
            rows.append(dict(ticker=tk, qty=sum(t["qty"] for t in trs), inv=inv_t, px=p,
                             pnl=pnl_t, ret=(pnl_t / inv_t if inv_t else None),
                             exc=pnl_t - bm_t, warn=(tk in s["warn_split"])))
        rows.sort(key=lambda r: -r["exc"])
        s["rows"] = rows
    return out


# ── 4) SVG (파이썬에서 그린다 — 조각은 스크립트를 못 싣는다) ─────────────────
def svg_lines(series, labels=None, w=760, h=232, pad=46, zero=None):
    """series = [(이름, [(x라벨, y)…][, 색[, 점선]])…] → 마우스 대응 가능한 SVG.

    🚨 그림은 파이썬이 그리고 **호버는 JS 가 얹는다**(FRAG_SCRIPT.chart 배선). 그래서 점
      좌표·라벨을 data-ch 에 JSON 으로 함께 싣는다 — JS 가 화면에서 다시 계산하면 선과
      툴팁이 서로 다른 수를 말하는 날이 온다(이 저장소가 되풀이 밟은 «경로가 둘» 유형).
    ⚠ x 도메인은 전 계열 라벨의 합집합이다. 첫 계열만 세면 짧은 계열(예상 꼬리)이 캔버스
      밖으로 나간다 — 실측 868>760.
    """
    if not series or not series[0][1]:
        return ""
    series = [(t + (None, False)[len(t) - 2:]) if len(t) < 4 else t for t in series]
    ys = [y for t in series for _x, y in t[1]]
    lo, hi = min(ys), max(ys)
    if hi - lo < 1e-9:
        hi = lo + 1.0
    span = hi - lo
    lo, hi = lo - span * 0.06, hi + span * 0.06        # 위아래 여백 — 선이 테두리에 붙지 않게
    xi = {}
    for t in series:
        for lb, _v in t[1]:
            if lb not in xi:
                xi[lb] = len(xi)
    n = len(xi)
    labs = [lb for lb, _i in sorted(xi.items(), key=lambda kv: kv[1])]
    colors = ["var(--accent)", "var(--champ)", "var(--rp)", "var(--hot)", "var(--deploy)"]

    def X(i):
        return pad + (w - pad - 12) * (i / max(1, n - 1))

    def Y(v):
        return (h - 26) - (h - 46) * ((v - lo) / (hi - lo))

    parts = ['<svg viewBox="0 0 %d %d" role="img" class="ichart" preserveAspectRatio="none" '
             'style="width:100%%;height:auto">' % (w, h)]
    # 가로 격자 4줄 + 값 라벨 — 눈금이 없으면 곡선의 기울기를 눈으로만 재게 된다
    for k in range(5):
        v = lo + (hi - lo) * k / 4.0
        yy = Y(v)
        parts.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" stroke="var(--line-soft)" stroke-width="1"/>'
                     % (pad, yy, w - 12, yy))
        parts.append('<text x="%d" y="%.1f" font-size="10" text-anchor="end" fill="var(--muted)" '
                     'font-family="var(--mono)">%.1f</text>' % (pad - 5, yy + 3.5, v))
    if lo < 100 < hi:      # 기준선(=100) 은 굵게 — 연초 후 그림에서 손익분기다
        parts.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" stroke="var(--line)" '
                     'stroke-dasharray="3 3"/>' % (pad, Y(100), w - 12, Y(100)))
    # 🚨 초과수익 그림에서는 **0선이 곧 벤치마크**다(0 = 지수와 같음). 그래서 벤치마크와
    #   같은 색으로 긋는다(2026-08-21 사용자 지시 «전략 차트에도 마찬가지로 빨간색»).
    #   ⚠ 연초 후 그림의 100선은 시작점일 뿐 벤치가 아니라서 이 색을 안 준다 — 그래서
    #     자동으로 칠하지 않고 부르는 쪽이 zero 를 넘길 때만 칠한다.
    if zero and lo < 0 < hi:
        parts.append('<line x1="%d" y1="%.1f" x2="%d" y2="%.1f" stroke="%s" stroke-width="1.4" '
                     'stroke-dasharray="5 3"/>' % (pad, Y(0), w - 12, Y(0), zero))
    # 세로 눈금 — 월이 바뀌는 자리에만(날짜를 다 적으면 겹친다)
    _pm = ""
    for i, lb in enumerate(labs):
        mm = str(lb)[:7]
        if mm != _pm and i:
            _pm = mm
            parts.append('<line x1="%.1f" y1="%d" x2="%.1f" y2="%.1f" stroke="var(--line-soft)" '
                         'stroke-width="1"/>' % (X(i), 12, X(i), h - 26))
            parts.append('<text x="%.1f" y="%d" font-size="9.5" text-anchor="middle" fill="var(--muted)" '
                         'font-family="var(--mono)">%s</text>' % (X(i), h - 12, esc(mm[5:] + "월")))
        elif not i:
            _pm = mm
    for k, t in enumerate(series):
        _name, pts, color, dash = t[0], t[1], t[2] or colors[k % len(colors)], t[3]
        d = " ".join("%s%.1f,%.1f" % ("M" if i == 0 else "L", X(xi[lb]), Y(v))
                     for i, (lb, v) in enumerate(pts))
        parts.append('<path d="%s" fill="none" stroke="%s" stroke-width="1.8" stroke-linejoin="round"%s/>'
                     % (d, color, ' stroke-dasharray="4 3"' if dash else ''))
    # 호버 장치 — 세로 안내선과 점은 JS 가 움직인다(처음엔 숨김)
    parts.append('<line class="chguide" x1="0" y1="12" x2="0" y2="%d" stroke="var(--muted)" '
                 'stroke-width="1" opacity="0"/>' % (h - 26))
    for k, t in enumerate(series):
        parts.append('<circle class="chdot" r="3.2" fill="%s" opacity="0"/>'
                     % (t[2] or colors[k % len(colors)]))
    parts.append('</svg>')
    if labels:
        parts.append('<div class="chlgd">' + "".join(
            '<span><i style="background:%s"></i>%s</span>' % (colors[k % len(colors)], esc(lb))
            for k, lb in enumerate(labels)) + "</div>")
    # JS 가 읽는 자료 — 라벨 목록 · 계열별 (이름, 색, 라벨→값)
    import json as _json
    meta = {"pad": pad, "w": w, "h": h, "lo": lo, "hi": hi, "labs": labs,
            "s": [{"n": t[0], "c": (t[2] or colors[k % len(colors)]),
                   "v": [[lb, round(v, 4)] for lb, v in t[1]]}
                  for k, t in enumerate(series)]}
    return ('<div class="chwrap" data-ch="%s">%s<div class="chtip" hidden></div></div>'
            % (esc(_json.dumps(meta, ensure_ascii=False, separators=(",", ":"))), "".join(parts)))


# ── 5) 렌더 ──────────────────────────────────────────────────────────────────
def render_fund(fund, idx, slug, label, nav, fx, hold, cons, trades, px, lvl, axis):
    H = []
    asof = max(hold[fund])
    rows = hold[fund][asof]
    # 시점 규약(사용자 확인 + 실측 검증 2026-08-20): 기준가(D) = 미국 D-1(직전 거래일) 종가
    # × D일 한국마감 환율. 그래서 NAV·환율은 보유 기준일로 «자르지 않고» 시트의 최신 행을
    # 쓴다 — 화면의 미국 종가(최신)와 짝이 맞는 것은 최신 NAV 다. 종전에는 last_leq(보유일)로
    # 잘라 하루 낡은 NAV 를 최신 종가 옆에 붙였다.
    nav_d = max(nav[fund])
    nav_v, base_p = nav[fund][nav_d]
    fx_d = max(fx)
    fx_v = fx[fx_d]
    _fxh_d, fx_hold = last_leq(fx, asof)   # 보유 원장과 같은 날 환율 — 원장 평가액과의 검산용

    by_asset = {}
    for r in rows:
        by_asset[r["asset"]] = by_asset.get(r["asset"], 0.0) + r["val_krw"]
    stocks = [r for r in rows if r["asset"] == "1"]
    sleeve = sum(r["val_krw"] for r in stocks) or 1.0
    w_stk = by_asset.get("1", 0) / nav_v
    w_etf = by_asset.get("3", 0) / nav_v
    w_fut = by_asset.get("4", 0) / nav_v          # 노셔널 — NAV 구성이 아니라 노출
    w_cash = (by_asset.get("5", 0) + by_asset.get("B", 0)) / nav_v

    # 연초 후 — 기준가 vs 지수(원화환산), 시작일 = 100.
    # 🚨 지수는 last_lt(엄격히 이전 날) — T-1 짝맞춤이다. 실측(2026-08-20): 기준가 일수익이
    #   같은날 지수와는 크게 어긋나고(08-19: −2.85% vs −1.22%) 전일 지수와 맞는다(−2.66%).
    #   last_leq 로 짝지으면 벤치 곡선이 펀드보다 하루 앞서 달려 꼬리에서 가짜 괴리가 생긴다.
    #   주말·휴일의 lvl 패딩(전일 값 복제)이 last_lt 와 만나면 자연스럽게 «직전 미국 거래일
    #   종가»가 된다 — 월요일 기준가(D)의 짝은 일요일 패딩 = 금요일 종가.
    lvl_i = lvl[idx]
    nav_ser = sorted(nav[fund].items())
    # ── 연초 후의 0점 (2026-08-21 사용자 검증으로 고침) ──────────────────────
    # 🚨 종전에는 시트 첫 행(전년 12-31 기준가)을 0점으로 썼다. 그런데 T-1 규약 때문에
    #   12-31 기준가는 **미국 12-30 종가**를 담는다 — 미국 12-31 하루(전년 몫: 2025-12-31
    #   SPX −0.74% · NDX −0.84%)가 새해 수익률로 새어 들어왔다. 그래서 2A81 연초후가
    #   사내 화면 +9.6% 대비 +8.67% 로 낮았다(차이 ≈ 정확히 그 하루 + 하루 지연).
    # → 0점 = **새해 첫 영업일의 기준가**(보통 1/2 · 미국 12-31 종가 반영). 사내 잣대와 같다.
    #   ⚠ «값이 달라진 첫 행» 으로 찾으면 안 된다 — 1/1 휴일 행도 이자 발생분만큼 미세하게
    #     달라서(실측 +0.00006%) 그 규칙이 1/1 을 잘못 문다. 시트는 달력일 전 행을 담고
    #     주말·휴일은 복제라, «1/2 이후 첫 평일» 이 새해 첫 기준가다(설날은 1월 말이라 안전).
    # ⚠ 시트가 연중에 시작하면(설정 첫해) 첫 행이 그대로 0점이다 — 그때는 새어 들 것이 없다.
    _yr = nav_ser[-1][0][:4]
    d0 = nav_ser[0][0]
    if d0[:4] != _yr:
        for _d2, _v2 in nav_ser:
            if _d2 >= _yr + "-01-02" and dt.date.fromisoformat(_d2).weekday() < 5:
                d0 = _d2
                break
        # 0점 이전 행(전년말·1/1 복제)은 곡선에서도 뺀다 — 카드(연초후)와 곡선이 같은
        # 0점을 써야 한다. 안 빼면 곡선 시작 100 이 카드와 다른 날을 가리킨다.
        nav_ser = [(_d2, _v2) for _d2, _v2 in nav_ser if _d2 >= d0]
    bp0 = nav[fund][d0][1]
    _i0d, i0 = last_lt(lvl_i, d0)
    _f0d, f0 = last_leq(fx, d0)
    fund_pts, bm_pts = [], []
    for d, (_n, bp) in nav_ser:
        _id, iv = last_lt(lvl_i, d)
        _fd, fv = last_leq(fx, d)
        if iv and fv and i0 and f0:
            fund_pts.append((d, bp / bp0 * 100))
            bm_pts.append((d, (iv * fv) / (i0 * f0) * 100))
    # ── 예상 꼬리 (2026-08-20 사용자 지시) ──────────────────────────────────
    # «차트 최근 1일(아직 NAV 에 반영 안 된 부분)은 빨간색으로 표시해줘. 예상이라는 의미로.»
    # 마지막 기준가가 담은 미국 종가(paired)보다 새 미국 종가(asof_us 쪽)가 있으면,
    # 그 구간의 지수 수익률을 기준가에 얹어 **추정** 점을 만든다 — 펀드는 지수를 복제하므로
    # 근사가 서지만, 전략 틸트·비용이 빠진 값이라 «예상» 이라 부르고 점선·빨강으로 가른다.
    # ⚠ 환율은 최신 시트 값(fx_v)을 그대로 쓴다 — 다음 날 한국마감 환율은 아직 없다.
    est_f, est_b = [], []
    if fund_pts:
        _paired_d, _paired_lvl = last_lt(lvl_i, nav_ser[-1][0])
        _new = [(d2, v2) for d2, v2 in sorted(lvl_i.items())
                if d2 > _paired_d and v2 is not None and v2 != _paired_lvl]
        # 패딩(전일 복제) 행은 값이 같아 걸러진다 — 남는 것이 진짜 새 종가다
        if _new and _paired_lvl:
            _ed, _ev = _new[-1]
            _fxr = fx_v / (last_leq(fx, nav_ser[-1][0])[1] or fx_v)
            est_f = [fund_pts[-1], (_ed, fund_pts[-1][1] * (_ev / _paired_lvl) * _fxr)]
            est_b = [bm_pts[-1], (_ed, (_ev * fx_v) / (i0 * f0) * 100)]
    ytd_f = fund_pts[-1][1] / 100 - 1 if fund_pts else None
    ytd_b = bm_pts[-1][1] / 100 - 1 if bm_pts else None
    _e_f = (est_f[1][1] / 100 - 1) if est_f else None
    _e_b = (est_b[1][1] / 100 - 1) if est_b else None

    cons_d, cmap = cons[idx]
    my_trades = [t for t in trades if t["index"] == idx]
    asof_us = axis[-1]      # 최신 미국 거래일(패딩 제외 축의 끝) — 매매 평가·성과의 기준
    perf = strat_perf(my_trades, px, lvl_i, asof_us) if my_trades else {}
    tot_exc = sum(s["last"]["pnl"] - s["last"]["bm"] for s in perf.values() if s.get("last"))
    tot_bp = tot_exc * fx_v / nav_v * 1e4 if perf else 0.0

    # 첫 탭이 보인다 — 순서를 FUNDS 가 정하므로 여기 이름을 박으면 순서를 바꿀 때 둘이 갈린다.
    H.append('<section class="tabpane" id="pane-%s"%s>' % (slug, "" if slug == FUNDS[0][2] else " hidden"))
    H.append('<div class="fhead"><h2>%s <span class="fcode">%s · %s</span></h2>'
             '<div class="asofline">보유 %s · NAV·기준가 %s · 환율 %s (%s) · 미국 종가 %s · 지수비중 %s'
             '<br>기준가(D) = 미국 D−1 종가 × D일 한국마감 환율 — 최신 종가와 짝은 최신 기준가</div></div>'
             % (esc(label), esc(fund), esc(idx), esc(asof), esc(nav_d), esc(fx_d), num(fx_v),
                esc(asof_us), esc(cons_d)))

    # ① 개요
    H.append('<h3>① 펀드 개요</h3><div class="cards">')
    cards = [
        ("순자산(NAV)", "%s억원" % num(nav_v / 1e8, 0), "기준가 " + num(base_p), 0),
        # 예상 반영값을 **부제에** 같이 적는다(2026-08-21 사용자 지시 «근사치 반영된 것도»).
        #   머리 숫자는 여전히 확정값이다 — 예상을 머리에 올리면 확정과 구별이 사라진다.
        ("연초 후(기준가)", pct(ytd_f, 2, True),
         ("지수 " + pct(ytd_b, 2, True) +
          ((" · 예상 " + pct(_e_f, 2, True)) if _e_f is not None else "")), ytd_f or 0),
        ("연초 후 초과", pct((ytd_f - ytd_b) if None not in (ytd_f, ytd_b) else None, 2, True),
         ("기준가 − 지수·원화" +
          ((" · 예상 " + pct(_e_f - _e_b, 2, True)) if None not in (_e_f, _e_b) else "")),
         (ytd_f - ytd_b) if None not in (ytd_f, ytd_b) else 0),
        # 🚨 전략 기여의 분자(초과손익)는 이미 **최신 미국 종가**로 잰다 — 기준가보다 하루
        #   앞선다. 분모(NAV·환율)만 시트 값이라, 이 카드는 이미 «근사치 반영»된 수다.
        #   그 사실을 부제가 말한다(종전에는 안 적어 확정값처럼 읽혔다).
        ("전략 NAV 기여", ("%+.1f bp" % tot_bp) if perf else "—",
         ("미국 %s 종가 기준 · NAV 는 %s" % (asof_us, nav_d)) if perf else "매매 없음", tot_bp),
    ]
    for ci, (k, v, sub, sign) in enumerate(cards):
        # cv/cs 의 id 는 남긴다 — 지금은 읽는 곳이 없지만 자리 표식이고, 있어도 해가 없다.
        H.append('<div class="card"><div class="ck">%s</div><div class="cv %s" id="cv-%s-%d">%s</div>'
                 '<div class="cs" id="cs-%s-%d">%s</div></div>'
                 % (esc(k), cls_sign(sign), slug, ci, esc(v), slug, ci, esc(sub)))
    H.append("</div>")
    # (2026-08-20 저녁) «최신 기준가 반영» 수동 입력 폼은 반나절 만에 걷었다 —
    #   사용자: «기준가·NAV·환율 넣는 건 다 빼줘. 번거로운 일이야.»
    #   대신 아래 차트가 **미반영 최근 거래일을 지수로 추정해** 빨간 선으로 자동 연장한다.
    # 🚨 2026-08-21 사용자 지시 «벤치마크 지수는 빨간색». 그래서 예상 꼬리를 빨강으로
    #   구별하던 방식은 못 쓴다 — 예상은 **각 계열 제 색 + 점선**으로만 가른다.
    #   (색은 «무엇의 선인가», 점선은 «확정인가 예상인가» — 축을 둘로 나누는 편이 옳다.)
    C_FUND, C_BM = "var(--accent)", "var(--hot)"
    _ser = [("펀드", fund_pts, C_FUND, False), ("지수", bm_pts, C_BM, False)]
    if est_f:
        _ser.append(("펀드 예상", est_f, C_FUND, True))
        _ser.append(("지수 예상", est_b, C_BM, True))
    H.append('<div class="chart" id="ytd-%s">%s%s</div>' % (slug, svg_lines(
        _ser, labels=["펀드 기준가", "%s 원화환산" % idx.split()[0]]),
        ('<p class="pnote" style="margin:6px 0 0">— <b>점선</b> = 아직 기준가에 반영 안 된 '
         '미국 %s 종가를 지수 수익률로 얹은 <b>예상</b>입니다(환율은 최신 시트 값 고정). '
         '전략 틸트·비용이 빠진 근사치입니다.</p>' % esc(est_f[1][0])) if est_f else ''))
    H.append('<table class="mini"><thead><tr><th>자산 구성</th><th class="tnum">NAV 대비</th></tr></thead><tbody>')
    for lbl2, w in (("개별주식", w_stk), ("지수 ETF", w_etf), ("현금·증거금", w_cash)):
        H.append('<tr><td>%s</td><td class="tnum">%s</td></tr>' % (esc(lbl2), pct(w, 1)))
    H.append('<tr><td>지수선물 노셔널(별도)</td><td class="tnum">%s</td></tr>' % pct(w_fut, 1))
    H.append('<tr><td><b>총 지수 노출</b></td><td class="tnum"><b>%s</b></td></tr>' % pct(w_stk + w_etf + w_fut, 1))
    H.append("</tbody></table>")

    # ② 보유 vs 지수
    held = {}
    for r in stocks:
        h = held.setdefault(r["ticker"], {"qty": 0.0, "val": 0.0, "px": r["px"], "name": r["name"]})
        h["qty"] += r["qty"]
        h["val"] += r["val_krw"]
    # 전략 매매의 종목별 순수량 — «펀드비중 = 패시브 복제 + 전략 틸트» 를 가르는 열쇠다.
    # 전략비중 = 순수량 × 원장 종가(USD) × 보유일 환율 ÷ 슬리브. 원장 내부 눈금(fx_hold)과
    # 같은 환율을 쓴다 — 최신 환율을 섞으면 패시브 = 펀드 − 전략 이 안 맞아떨어진다.
    strat_q = {}
    for t in my_trades:
        strat_q[t["ticker"]] = strat_q.get(t["ticker"], 0.0) + t["qty"]
    n_match = sum(1 for t in held if t in cmap)
    tbl = []
    for t, h in held.items():
        w_f = h["val"] / sleeve
        w_i = cmap.get(t, (0.0, None, None))[0]
        gics = cmap.get(t, (None, None, ""))[2] or ""
        # 패시브 수량 검산은 원장 «내부» 눈금으로 — 원장 평가액(원화)이 원장 종가×보유일 환율로
        # 만들어졌으니, 최신 환율을 섞으면 자기모순이 된다(수량 괴리가 환율 차이로 오염).
        p_qty = (w_i * sleeve / (h["px"] * fx_hold)) if (h["px"] and fx_hold) else None
        # 🚨 2026-08-21 사용자 지적 «보유 수량 = 패시브 수량 + 전략 수량인데 숫자가 이상해».
        #   맞다 — 종전 「패시브수량」은 **지수비중으로 역산한 이론값**(p_qty)이라 보유수량과
        #   애초에 합이 안 맞는 다른 개념이었다. 항등식이 서려면 패시브를 «보유 − 전략» 으로
        #   도출해야 한다. 비중 열(wp = wf − ws)은 이미 그렇게 돼 있어 맞았고, 수량만 어긋나
        #   있었다 — 같은 표에서 두 열이 다른 정의를 쓰고 있었던 셈이다.
        #   이론값은 «지수기준» 이라는 제 이름으로 옮기고 괴리 계산에만 쓴다.
        _sq = strat_q.get(t, 0.0)
        w_s = (_sq * h["px"] * fx_hold / sleeve) if (h["px"] and fx_hold and _sq) else 0.0
        tbl.append(dict(t=t, name=h["name"], gics=gics, wi=w_i, wf=w_f, d=(w_f - w_i) * 100,
                        ws=w_s, wp=w_f - w_s,
                        qty=h["qty"], sq=_sq, pq_real=h["qty"] - _sq,
                        pq=p_qty, dq=(h["qty"] - p_qty) if p_qty is not None else None))
    tbl.sort(key=lambda r: -r["wf"])
    # 문턱 없음(2026-08-20 사용자 지시 «비중 상관없이») — 지수에 있는데 안 든 것 전부.
    only_idx = sorted(((t, v[0], v[1]) for t, v in cmap.items() if t not in held),
                      key=lambda x: -x[1])
    off_idx = [r for r in tbl if r["t"] not in cmap]     # 보유 중인데 지수에 없는 것

    H.append('<h3>② 보유 vs 지수 <span class="hnote">주식 슬리브 %d종 · 지수 %d종%s</span></h3>'
             % (len(held), len(cmap),
                (' · <b>지수 밖 보유 %d종</b>' % len(off_idx)) if off_idx else ''))
    H.append('<div class="filterbar"><input type="search" class="rowfilter" data-target="tb-%s" '
             'placeholder="티커·이름으로 거르기" aria-label="표 필터"></div>' % slug)
    H.append('<div class="tblwrap tall"><table class="big" id="tb-%s"><thead><tr>'
             '<th>티커</th><th>이름</th><th>섹터</th><th class="tnum">지수비중</th><th class="tnum">펀드비중</th>'
             '<th class="tnum">패시브</th><th class="tnum">전략</th>'
             '<th class="tnum">차이(%%p)</th>'
             '<th class="tnum">보유 수량</th><th class="tnum">패시브 수량</th><th class="tnum">전략 수량</th>'
             '<th class="tnum">지수기준</th><th class="tnum">괴리</th>'
             '</tr></thead><tbody>' % slug)
    for r in tbl:
        # 지수 밖 보유 — 편출 뒤 잔존 보유 등. 지수비중 0.00 만으로는 «아주 작다» 와
        # «지수에 없다» 가 안 갈린다. 배지로 말한다.
        _off = ' <span class="offb" title="지수 구성종목이 아니다">밖</span>' if r["t"] not in cmap else ''
        # 펀드비중 칸 음영 — 초록 = 오버웨이트 · 빨강 = 언더웨이트. 진하기는 |차이| 에 비례
        # (0.5%p 에서 상한 30%). 색만으로 안 가르고 차이 열이 수치를 그대로 말한다.
        # 🚨 완성된 조각(_off·_bg)을 **포맷 문자열에 이어 붙이지 않는다.** 붙였다가
        #   color-mix 의 «30%,» 가 서식 문자로 읽혀 빌드가 죽었다(ValueError, 2026-08-21).
        #   조각은 값으로만 넘긴다 — 그러면 그 안에 무엇이 들어 있든 안전하다.
        _shade = min(30.0, abs(r["d"]) * 60.0)
        _bg = (' style="background:color-mix(in srgb,var(--%s) %d%%,transparent)"'
               % ("good" if r["d"] > 0 else "hot", round(_shade))) if abs(r["d"]) >= 0.005 else ''
        H.append('<tr><td class="tk">%s%s</td><td>%s</td><td class="sec">%s</td>'
                 '<td class="tnum">%s</td><td class="tnum"%s>%s</td>'
                 '<td class="tnum">%s</td><td class="tnum %s">%s</td>'
                 '<td class="tnum %s">%+.2f</td>'
                 '<td class="tnum"><b>%s</b></td><td class="tnum">%s</td><td class="tnum %s">%s</td>'
                 '<td class="tnum sub">%s</td><td class="tnum %s">%s</td></tr>'
                 % (esc(r["t"]), _off, esc(r["name"][:26]), esc((r["gics"] or "")[:16]),
                    pct(r["wi"], 2), _bg, pct(r["wf"], 2),
                    pct(r["wp"], 2), ("tk" if r["ws"] else ""), (pct(r["ws"], 2) if r["ws"] else "—"),
                    cls_sign(r["d"]), r["d"],
                    num(r["qty"], 0), num(r["pq_real"], 0),
                    ("tk" if r["sq"] else ""), (num(r["sq"], 0) if r["sq"] else "—"),
                    num(r["pq"], 0) if r["pq"] is not None else "—",
                    cls_sign(r["dq"] or 0), ("%+d" % round(r["dq"])) if r["dq"] is not None else "—"))
    H.append("</tbody></table></div>")
    H.append('<p class="pnote" style="margin:6px 0 12px">'
             '<b>보유 수량 = 패시브 수량 + 전략 수량</b>입니다 — 전략 수량은 매매 원장의 종목별 '
             '순수량이고, 패시브는 그 나머지(지수 복제분)입니다. 비중 열도 같은 분해입니다. '
             '<b>지수기준</b>은 «지수비중 × 주식슬리브 ÷ (종가×환율)» 로 역산한 이론 수량이고, '
             '<b>괴리</b> = 보유 − 지수기준 입니다 — 액티브 틸트의 수량 표현이라 위 분해와는 '
             '다른 질문에 답합니다.</p>')
    if off_idx:
        H.append('<details open><summary>보유 중인데 지수에 없는 %d종 — 편출 뒤 잔존 등</summary>'
                 '<div class="tblwrap"><table class="mini"><tbody>' % len(off_idx))
        for r in off_idx:
            H.append('<tr><td class="tk">%s</td><td>%s</td><td class="tnum">%s</td></tr>'
                     % (esc(r["t"]), esc(r["name"][:30]), pct(r["wf"], 2)))
        H.append("</tbody></table></div></details>")
    if only_idx:
        H.append('<details><summary>지수에 있는데 미보유 %d종 — 전부(문턱 없음)</summary>'
                 '<div class="tblwrap"><table class="mini"><tbody>' % len(only_idx))
        for t, w, nm in only_idx:
            H.append('<tr><td class="tk">%s</td><td>%s</td><td class="tnum">%s</td></tr>'
                     % (esc(t), esc((nm or "")[:30]), pct(w, 2)))
        H.append("</tbody></table></div></details>")

    # ③④ 는 웹 앱(js/portfolio_app.js)이 그린다 — DB 원장 + 웹 입력 원장을 합쳐 라이브로
    #   계산해야 하기 때문이다(웹 입력은 이 생성기가 돌 때 존재하지 않는다). 파이썬 계산은
    #   아래 <details> 정적 스냅샷으로 남겨 두 엔진의 교차검증 대상이 된다 — PF.check 에도
    #   수치를 실어 앱이 기계 대조한다(불일치면 화면에 ⚠).
    H.append('<h3>③ 매매 원장 <span class="hnote">DB(mp)+웹 입력 통합 · 체결가 = 당시 종가(실제 체결가 아님)</span></h3>')
    H.append('<div class="appbox" id="ledger-%s"><p class="jswait">웹 앱이 그립니다 — 이 문구가 남아 있으면 js/portfolio_app.js 로드 실패</p></div>' % slug)
    H.append('<h3>④ 전략 성과·기여 <span class="hnote">라이브 계산 · BM = 같은 날 같은 금액을 지수에(매매 시점 일치)</span></h3>')
    H.append('<div class="appbox" id="perf-%s"><p class="jswait">웹 앱이 그립니다…</p></div>' % slug)

    H.append('<details class="dbstatic"><summary>정적 스냅샷 — 생성 시점 파이썬 계산(교차검증용)</summary>')
    H.append('<h4>매매 내역(DB 원장만)</h4>')
    if my_trades:
        H.append('<div class="tblwrap"><table class="big"><thead><tr><th>일자</th><th>전략</th><th>티커</th>'
                 '<th class="tnum">수량</th><th class="tnum">체결가</th><th class="tnum">금액(USD)</th>'
                 '<th class="tnum">현재가</th><th class="tnum">평가손익</th></tr></thead><tbody>')
        for t in my_trades:
            _pd, p = last_leq(px.get(t["ticker"], {}), asof_us)
            pnl = t["qty"] * (p - t["px"]) if p is not None else None
            H.append('<tr%s><td>%s</td><td>%s</td><td class="tk">%s</td><td class="tnum">%s</td>'
                     '<td class="tnum">%s</td><td class="tnum">%s</td><td class="tnum">%s</td>'
                     '<td class="tnum %s">%s</td></tr>'
                     % (' class="warnrow"' if t["qty"] == 0 else "",
                        esc(t["dt"]), esc(t["strategy"]), esc(t["ticker"]), num(t["qty"], 0),
                        num(t["px"]), num(t["qty"] * t["px"], 0), num(p) if p is not None else "—",
                        cls_sign(pnl or 0), num(pnl, 0) if pnl is not None else "—"))
        H.append("</tbody></table></div>")
        zeros = [t for t in my_trades if t["qty"] == 0]
        if zeros:
            H.append('<p class="warn">⚠ 수량 0 인 행 %d건 — 매매가 아니라 입력 흔적이다. 원장에서 지우는 게 맞다.</p>' % len(zeros))
    else:
        H.append("<p>이 지수에는 아직 전략 매매가 없다.</p>")

    # ④ 성과·기여 — 정적 스냅샷 쪽
    H.append('<h4>전략 성과·기여(DB 원장만)</h4>')
    if perf:
        H.append('<div class="tblwrap"><table class="big"><thead><tr><th>전략</th><th class="tnum">매수원금(USD)</th>'
                 '<th class="tnum">손익</th><th class="tnum">수익률</th><th class="tnum">BM</th>'
                 '<th class="tnum">초과</th><th class="tnum">NAV 기여(bp)</th></tr></thead><tbody>')
        for sname, s in sorted(perf.items()):
            L = s.get("last") or {}
            exc = L.get("pnl", 0) - L.get("bm", 0)
            contrib = exc * fx_v / nav_v * 1e4
            wmark = (" ⚠분할의심 " + ",".join(s["warn_split"])) if s["warn_split"] else ""
            H.append('<tr><td>%s%s</td><td class="tnum">%s</td><td class="tnum %s">%s</td>'
                     '<td class="tnum">%s</td><td class="tnum">%s</td><td class="tnum %s">%s</td>'
                     '<td class="tnum %s">%+.1f</td></tr>'
                     % (esc(sname), esc(wmark), num(L.get("inv"), 0), cls_sign(L.get("pnl")), num(L.get("pnl"), 0),
                        pct(L.get("ret"), 2, True), pct(L.get("bm_ret"), 2, True),
                        cls_sign(exc), num(exc, 0), cls_sign(contrib), contrib))
        H.append("</tbody></table></div>")
        series = []
        for sname, s in sorted(perf.items()):
            pts = [(d, (pnl - bm) / inv * 100 if inv else 0.0) for d, pnl, bm, inv in s["curve"]]
            if pts:
                series.append((sname, pts))
        if series:
            H.append('<div class="chart"><div class="chtitle">전략별 누적 초과수익(%%, 매수원금 대비)'
                     '<span class="hnote">빨간 0선 = 지수와 같은 성과</span></div>%s</div>'
                     % svg_lines(series, labels=[n[:16] for n, _p in series], zero="var(--hot)"))
        for sname, s in sorted(perf.items()):
            if not s.get("rows"):
                continue
            H.append('<details><summary>%s — 종목별 분해</summary><div class="tblwrap"><table class="mini"><thead><tr>'
                     '<th>티커</th><th class="tnum">순수량</th><th class="tnum">현재가</th><th class="tnum">손익(USD)</th>'
                     '<th class="tnum">수익률</th><th class="tnum">초과(USD)</th></tr></thead><tbody>' % esc(sname))
            for r in s["rows"]:
                H.append('<tr><td class="tk">%s%s</td><td class="tnum">%s</td><td class="tnum">%s</td>'
                         '<td class="tnum %s">%s</td><td class="tnum">%s</td><td class="tnum %s">%s</td></tr>'
                         % (esc(r["ticker"]), " ⚠" if r["warn"] else "", num(r["qty"], 0), num(r["px"]),
                            cls_sign(r["pnl"]), num(r["pnl"], 0), pct(r["ret"], 2, True),
                            cls_sign(r["exc"]), num(r["exc"], 0)))
            H.append("</tbody></table></div></details>")
    else:
        H.append("<p>전략 매매가 없어 성과를 계산할 것이 없다.</p>")
    H.append("</details>")

    H.append("</section>")
    # 웹 앱 교차검증용 — 파이썬이 계산한 전략별 (매수원금, 손익, BM손익). JS 엔진이 같은
    # 원장(DB분)으로 같은 수를 내는지 화면에서 기계 대조한다.
    perf_check = {sname: {"inv": round(s["last"]["inv"], 2), "pnl": round(s["last"]["pnl"], 2),
                          "bm": round(s["last"]["bm"], 2)}
                  for sname, s in perf.items() if s.get("last")}
    fmeta = dict(fund=fund, idx=idx, label=label, asof=asof, asof_us=asof_us,
                 nav=nav_v, base=base_p, fx=fx_v, sleeve=round(sleeve, 0), cons_d=cons_d,
                 check=perf_check,
                 cons={t: [round(v[0], 6), (v[1] or "")[:26], (v[2] or "")[:16]] for t, v in cmap.items()},
                 held={t: round(h["qty"], 2) for t, h in held.items()})
    # ── 배치 재정렬 (2026-08-20 사용자 지시) ─────────────────────────────────
    # «차트가 위쪽에 배치되게 하고 보유 vs 지수 이런 큰 테이블은 아래쪽에.»
    # 위 코드는 계산 의존 순서(개요 수치 → 표)대로 쌓는다 — 그 순서는 유지하고,
    # **표시 순서만** 여기서 바꾼다: 차트 → ①개요 → 매매 → 성과 → 보유 vs 지수.
    # 조립 코드를 통째로 재배열하지 않는 이유: 수치 계산이 문서 순서에 끼어 있어
    # 옮기다 참조가 깨진다(fund_pts 를 카드가 먼저 쓴다). 문자열 재배열이 안전하다.
    pane = "\n".join(H)
    import re as _re
    _h1 = pane.index('<h3>① 펀드 개요')
    _h2 = pane.index('<h3>② 보유 vs 지수')
    _h3 = pane.index('<h3>③ 매매 원장')
    _h4 = pane.index('<h3>④ 전략 성과·기여')
    _tail = pane.rindex('</section>')
    head, sec1, sec2 = pane[:_h1], pane[_h1:_h2], pane[_h2:_h3]
    sec_led, sec_perf = pane[_h3:_h4], pane[_h4:_tail]
    # ①에서 차트만 떼어 머리 바로 뒤로 — 자산구성 표는 개요에 남긴다
    _m = _re.search(r'<div class="chart" id="ytd-[^"]+">.*?</div>', sec1, _re.S)
    chart, sec1 = _m.group(0), sec1[:_m.start()] + sec1[_m.end():]
    # 번호는 표시 순서를 따른다 — 안 바꾸면 ①③④② 로 읽혀 빠진 줄이 있는 줄 안다
    # 2026-08-21 사용자 지시 «매매원장을 가장 아래로» — 표시 순서:
    #   차트 → ① 개요 → ② 성과·기여 → ③ 보유 vs 지수 → ④ 매매 원장
    # ⚠ 정적 스냅샷(<details>)은 성과 절 끝에 붙어 있어 함께 따라간다 — 그게 맞다.
    sec_perf = sec_perf.replace('<h3>④ 전략 성과·기여', '<h3>② 전략 성과·기여', 1)
    sec2 = sec2.replace('<h3>② 보유 vs 지수', '<h3>③ 보유 vs 지수', 1)
    sec_led = sec_led.replace('<h3>③ 매매 원장', '<h3>④ 매매 원장', 1)
    pane = head + chart + sec1 + sec_perf + sec2 + sec_led + pane[_tail:]
    return pane, dict(fund=fund, asof=asof, nav_d=nav_d, sleeve_ratio=w_stk,
                              n_stocks=len(held), n_match=n_match, n_cons=len(cmap), meta=fmeta)



FRAG_SCRIPT = """<script>
(function(){
  var tabs=document.querySelectorAll('#content .tb');
  function show(id){
    tabs.forEach(function(x){x.setAttribute('aria-selected', x.dataset.tab===id?'true':'false');});
    document.querySelectorAll('#content .tabpane').forEach(function(pn){pn.hidden=(pn.id!=='pane-'+id);});
  }
  tabs.forEach(function(t){t.addEventListener('click',function(){show(t.dataset.tab);});});
  // 차트 호버 — 정적 SVG 위에 안내선·점·툴팁을 얹는다. 좌표는 그릴 때 실은 data-ch 를
  // 그대로 쓴다(화면이 다시 계산하면 선과 툴팁이 다른 수를 말한다).
  function wireCharts(root){
    (root || document).querySelectorAll('.chwrap[data-ch]').forEach(function(wrap){
      if(wrap.dataset.wired) return; wrap.dataset.wired='1';
      var M; try{ M=JSON.parse(wrap.getAttribute('data-ch')); }catch(e){ return; }
      var svg=wrap.querySelector('svg'), tip=wrap.querySelector('.chtip');
      var guide=wrap.querySelector('.chguide'), dots=wrap.querySelectorAll('.chdot');
      var n=M.labs.length, X=function(i){return M.pad+(M.w-M.pad-12)*(i/Math.max(1,n-1));};
      var Y=function(v){return (M.h-26)-(M.h-46)*((v-M.lo)/(M.hi-M.lo));};
      var byLab=M.s.map(function(se){var m={};se.v.forEach(function(p){m[p[0]]=p[1];});return m;});
      function hide(){ tip.hidden=true; guide.setAttribute('opacity','0');
        dots.forEach(function(d){d.setAttribute('opacity','0');}); }
      svg.addEventListener('mousemove',function(ev){
        var r=svg.getBoundingClientRect();
        var vx=(ev.clientX-r.left)/r.width*M.w;            // viewBox 좌표로 되돌린다
        var i=Math.round((vx-M.pad)/((M.w-M.pad-12)/Math.max(1,n-1)));
        i=Math.max(0,Math.min(n-1,i));
        var lab=M.labs[i];
        guide.setAttribute('x1',X(i)); guide.setAttribute('x2',X(i)); guide.setAttribute('opacity','.5');
        var rows='';
        M.s.forEach(function(se,k){
          var v=byLab[k][lab];
          if(v==null){ dots[k].setAttribute('opacity','0'); return; }
          dots[k].setAttribute('cx',X(i)); dots[k].setAttribute('cy',Y(v)); dots[k].setAttribute('opacity','1');
          rows+='<div><i style="background:'+se.c+'"></i>'+se.n+'<b>'+v.toFixed(2)+'</b></div>';
        });
        if(!rows){ hide(); return; }
        tip.innerHTML='<div class="chtd">'+lab+'</div>'+rows;
        tip.hidden=false;
        var px=(X(i)/M.w)*r.width;                          // 화면 픽셀로
        tip.style.left=Math.max(0,Math.min(r.width-tip.offsetWidth-2,px+10))+'px';
      });
      svg.addEventListener('mouseleave',hide);
    });
  }
  wireCharts();
  window.PFCHARTS = wireCharts;     // 웹 앱이 그린 차트도 같은 배선을 쓴다
  document.querySelectorAll('#content .rowfilter').forEach(function(inp){
    inp.addEventListener('input',function(){
      var q=inp.value.trim().toUpperCase();
      var tb=document.getElementById(inp.dataset.target);
      if(!tb)return;
      tb.querySelectorAll('tbody tr').forEach(function(tr){
        tr.hidden=!!q&&tr.textContent.toUpperCase().indexOf(q)<0;
      });
    });
  });
})();
</script>"""

NOTES = """<div class="notes"><h3>정의·한계</h3><ul>
<li><b>시점 규약</b> — 기준가(D)는 «미국 D−1(직전 거래일) 종가 × D일 한국마감 환율»을 담는다(실측 검증: 기준가 일수익은 전일 지수와 맞는다). 그래서 화면은 NAV·환율을 시트의 최신 행으로, 미국 종가를 최신 거래일로 각각 표시하고, 연초 후 차트는 지수를 하루 밀어(T-1) 기준가와 짝을 맞춘다.</li>
<li><b>펀드비중</b>은 종목 평가액(원화)을 개별주식 슬리브 합으로 나눈 것이다 — 지수비중과 같은 눈금이 되도록 주식 안에서 정규화했다. 펀드는 주식+ETF+선물로 지수를 복제하므로 NAV 분모로 보면 전 종목이 일괄 언더웨이트로 보인다.</li>
<li><b>패시브 수량</b> = 지수비중 × 주식슬리브(원) ÷ (종가 × USD환율). 괴리 = 실제 − 패시브. 액티브 틸트의 수량 표현이다.</li>
<li><b>체결가는 당시 종가다</b> — 실제 체결가가 아니다(원장을 쓰는 MP 엑셀이 종가를 박는다). 수익률·기여는 그만큼 근사치다.</li>
<li><b>BM 대조</b>는 같은 날 같은 금액을 지수에 넣었을 때의 손익이다(매매 시점 일치). 초과 = 전략 손익 − BM 손익. <b>NAV 기여(bp)</b> = 초과(USD) × 기준일 환율 ÷ NAV × 10,000.</li>
<li><b>배당 미반영</b> — 종가는 수정주가가 아니다. 보유 기간이 짧아 왜곡은 작지만 0이 아니다.</li>
<li>매매~기준일 사이 하루 ±40%를 넘는 가격변동이 있으면 <b>분할 의심 ⚠</b>를 단다 — 벤더가 분할을 소급 반영하지 않은 사고가 실측된 바 있다.</li>
<li><b>자료 원천</b>(2026-08-21 개편) — ① 사내 export 엑셀: NAV·환율·보유 원장. ② 사내 DB: <b>지수 구성종목·비중·GICS 하나만</b>. ③ 이 랩의 웹 자료(yfinance·매일 자동): 종목 종가·지수 레벨·분할 이력. ④ 웹 원장: 전략 매매. 갱신은 수동이고 상단의 생성 시각이 곧 이 화면의 기준이다.</li>
<li><b>분할 눈금</b> — 종가는 분할 소급 조정본이고, 원장 체결가도 같은 눈금으로 되맞춘다(가격÷분할비·수량×분할비, 금액 불변). 사내 종가는 분할을 소급하지 않아 그 자체가 틀린 값이었다.</li>
<li><b>지수 밖 보유</b>는 이 랩 유니버스(오늘의 S&amp;P 500 ∪ 나스닥 100)에 없어 가격이 빠진다 — 편출 예정이거나 합병 등으로 일시적으로 생긴 자리라 성과 계산에서 제외한다.</li>
<li><b>웹 원장</b>은 이 페이지에서 입력한 전략·매매다. 저장하면 열람 암호로 AES-256-GCM 암호화되어 저장소의 <span class="tk">data/portfolio_user.json</span> 에 커밋된다 — 평문은 저장소·서버 어디에도 남지 않고, git 이력이 곧 버전 관리라 언제든 과거 버전으로 되돌릴 수 있다. 쓰기 토큰은 이 기기 브라우저에만 저장된다.</li>
<li><b>웹 백테스트는 진단용이다.</b> 유니버스가 «현재» 구성종목이라 <b>생존편향</b>이 있고(랩 실측: 스타일에 따라 수~수십%p 부풀림), 지수는 PR·종가는 무배당이며, 창이 단일(패널 @PANEL@거래일)이라 통계적 유의성을 말할 수 없다. 배포 판단은 랩의 PIT 백테스트로만 한다.</li>
</ul></div>""".replace("@PANEL@", str(PANEL_DAYS))


def main() -> int:
    path, nav, fx, hold = load_xlsm()
    print("입력: %s" % os.path.basename(path))
    for f, _i, _s, _l in FUNDS:
        if f not in hold:
            raise SystemExit("해외 시트에 펀드 %s 가 없다 — export 범위를 확인할 것" % f)
        if f not in nav:
            raise SystemExit("NAV 시트에 펀드 %s 가 없다" % f)
    asof_by_fund = {f: max(hold[f]) for f, _i, _s, _l in FUNDS}
    held_tk = {r["ticker"] for f, _i, _s, _l in FUNDS
               for r in hold[f][max(hold[f])] if r["asset"] == "1" and r["ticker"]}
    cons, seed = load_db(asof_by_fund, held_tk)
    load_splits()
    # 유니버스 = 보유 ∪ 지수 구성 ∪ 씨앗 원장 티커
    uni = set(held_tk) | {t["ticker"] for t in seed}
    for _d, _cmap in cons.values():
        uni |= set(_cmap)
    px, lvl, axis = load_web(uni)
    # 🚨 씨앗 원장의 체결가를 **분할 소급 기준으로 되맞춘다.** 랩 종가가 조정본이라
    #   그대로 두면 분할 종목의 손익이 분할비만큼 틀린다(금액은 불변: 가격÷f · 수량×f).
    trades, _adj = [], []
    for t in seed:
        f = split_factor(t["ticker"], t["dt"])
        if f != 1.0:
            _adj.append("%s %s ×%.4g" % (t["ticker"], t["dt"], f))
            t = dict(t, px=t["px"] / f, qty=t["qty"] * f)
        trades.append(t)
    if _adj:
        print("  분할 되맞춤 %d건: %s" % (len(_adj), " · ".join(_adj[:6])))

    panes, checks = [], []
    for f, idx, slug, label in FUNDS:
        pane, chk = render_fund(f, idx, slug, label, nav, fx, hold, cons, trades, px, lvl, axis)
        panes.append(pane)
        checks.append(chk)

    # ── 실행 검산 — 화면에 내보내기 전에 여기서 막는다 ─────────────────────
    for c in checks:
        if not (0.3 <= c["sleeve_ratio"] <= 1.05):
            raise SystemExit("%s 주식슬리브/NAV=%.2f — 자산구분 해석이 깨졌다" % (c["fund"], c["sleeve_ratio"]))
        if c["n_match"] < c["n_stocks"] * 0.95:
            raise SystemExit("%s 보유 %d종 중 지수 매칭 %d — 티커 형식이 어긋났다"
                             % (c["fund"], c["n_stocks"], c["n_match"]))
        if c["nav_d"] != c["asof"]:
            print("⚠ %s NAV 기준일(%s) ≠ 보유 기준일(%s) — 이하 최근값으로 대체" % (c["fund"], c["nav_d"], c["asof"]))

    gen = dt.datetime.now().strftime("%Y-%m-%d %H:%M KST")
    tabs = "".join('<button class="tb" data-tab="%s" aria-selected="%s">%s · %s</button>'
                   % (slug, "true" if k == 0 else "false", label, f)
                   for k, (f, _i, slug, label) in enumerate(FUNDS))
    tabs += '<button class="tb" data-tab="bt" aria-selected="false">웹 백테스트</button>'

    # 웹 앱 데이터 블롭 — 원장 편집·성과 재계산·백테스트가 전부 이걸 먹는다.
    #   조각(=암호문) 안에만 들어간다. 실펀드 수치라 평문 JSON 으로 따로 두면 안 된다.
    panel = encode_panel(px, axis)
    asof_g = max(asof_by_fund.values())
    lvl_arr = {}
    for idx_name, key in (("NDX Index", "ndx"), ("SPX Index", "spx")):
        ser = lvl[idx_name]
        last = last_leq(ser, axis[0])[1]
        out = []
        for d in axis:
            last = ser.get(d, last)
            out.append(round(last, 2) if last is not None else None)
        lvl_arr[key] = out
    pf = {"gen": gen, "asof": asof_g, "dates": axis, "panel": panel, "lvl": lvl_arr,
          "funds": {slug: checks[k]["meta"] for k, (_f, _i, slug, _l) in enumerate(FUNDS)},
          "mp": [{"idx": t["index"], "dt": t["dt"], "s": t["strategy"],
                  "t": t["ticker"], "q": t["qty"], "p": t["px"]} for t in trades],
          "guard": SPLIT_GUARD}
    # "</script>" 조기 종결 방지 — JSON 문자열 값 안의 "</" 를 이스케이프한다.
    pf_json = json.dumps(pf, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")

    frag = ['<div class="pfhead"><div class="tabs" role="tablist">%s</div>'
            '<div id="pfsync" class="pfsync"></div>'
            '<div class="gen">생성 %s · 입력 %s</div></div>' % (tabs, esc(gen), esc(os.path.basename(path)))]
    frag.append('<div class="fundgrid">')
    frag += panes
    frag.append('</div>')
    frag.append('<section class="tabpane btpane" id="pane-bt" hidden>'
                '<h3>⑤ 웹 백테스트 <span class="hnote">동봉 종가 패널 %d거래일(%s~%s) · '
                '유니버스 = 현재 구성종목 — 생존편향 있음(아래 한계)</span></h3>'
                '<div class="appbox" id="btbox"><p class="jswait">웹 앱이 그립니다…</p></div>'
                '</section>' % (len(axis), esc(axis[0]), esc(axis[-1])))
    frag.append(NOTES)
    frag.append('<script>window.PF=%s;</script>' % pf_json)
    frag.append(FRAG_SCRIPT)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    body = "\n".join(frag)
    io.open(OUT, "w", encoding="utf-8", newline="").write(body)
    print("→ %s (%.0fKB)" % (os.path.relpath(OUT, ROOT), len(body.encode("utf-8")) / 1024))
    for c in checks:
        print("   %s: 보유 %s · 주식 %d종(지수 매칭 %d) · 슬리브/NAV %.1f%%"
              % (c["fund"], c["asof"], c["n_stocks"], c["n_match"], c["sleeve_ratio"] * 100))
    print("다음: python build/kb_lock.py --page portfolio   (열람 암호 입력 → portfolio.html 에 기록)")
    return 0


if __name__ == "__main__":
    import gate
    gate.run(main, "운용 포트폴리오")
